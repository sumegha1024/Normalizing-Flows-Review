import torch
import torch.optim as optim
import sys
import os
sys.path.append(os.path.join(sys.path[0],'Flows_scripts'))
import nn as nn_
import torch.nn as nn
from tqdm import tqdm
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import math
import scipy.stats as ss
import seaborn as sns
import random
import time
from torch.distributions import MultivariateNormal
import argparse
import pandas as pd
import flows
from utils import set_all_seeds
from sklearn import metrics
import statsmodels.api as sm


parser = argparse.ArgumentParser(description='inputs_regression')
parser.add_argument('--flows', default = 'NAF', type=str, help='type of flow model; Use None if result for flows is not required')
parser.add_argument('--mcmc', default = 'gibbs', type=str, help='type of mcmc method; Use None if result for mcmc is not required')
parser.add_argument('--rho',default =0.0, type=float, help='Correlation coefficient for design matrix')
parser.add_argument('--sigma',default =1.0, type=float, help='Standard dev for simulated data')
parser.add_argument('--tau',default =1.0, type=float, help='Standard dev for beta prior')
parser.add_argument('--sigma_prior', default=False, action='store_true', help='Whether we want sigma to be treated as constant or use a prior')
parser.add_argument('--lr', default = 0.0005, type=float, help='learning rate')
parser.add_argument('--seed',default =3, type=int, help='seed for simulation')
parser.add_argument('--out',default = '/mnt/home/premchan/Normalizing-Flows-Review/Temp/', type=str, help='path to results')
parser.add_argument('--data_dim',default =100, type=int, help='dimension of beta vector')
parser.add_argument('--n_data',default =100, type=int, help='number of samples for y')
parser.add_argument('--sparse',default =0.2, type=float, help='sparsity level in %')
parser.add_argument('--writecsv', default=False, action='store_true', help='Whether we want to write metrics to Result.xlsx')

args = parser.parse_args()

set_all_seeds(0)
#Dictionary of hyper-parameters - nsamps mh, epochs flows, epochs vi, hidden_dim for conditioner, thinning rate for gibbs 
tuning_dict= {'2_50':[101000,20000,20000,64,10],'2_100':[101000,20000,20000,64,10], '2_200':[101000,20000,20000,64,10], '20_50':[101000,20000,20000,64,10],'20_100':[101000,20000,20000,64,10],'20_200':[101000,30000,20000,64,10],
    '50_50':[101000,20000,100000,64,10],'50_100':[101000,20000,30000,64,10],'50_200':[101000,20000,20000,64,10],'100_50':[101000,30000,500000,128,10],'100_100':[101000,30000,250000,128,10],'100_200':[101000,30000,30000,128,10]}


#tuning_dict= {'2_50':[101000,20000,20000,64,10],'2_100':[101000,20000,20000,64,10], '2_200':[101000,20000,20000,64,10], '20_50':[101000,20000,20000,64,10],'20_100':[101000,20000,20000,64,10],'20_200':[101000,30000,20000,64,10],
#    '50_50':[101000,20000,500000,64,10],'50_100':[101000,20000,30000,64,10],'50_200':[101000,20000,20000,64,10],'100_50':[101000,30000,1000000,128,10],'100_100':[101000,30000,500000,128,10],'100_200':[101000,30000,30000,128,10]}

epochs_gibbs=tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][0]
epochs_flows =tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][1]
epochs_vi =tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][2]
cmade_dim =tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][3]
thin = tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][4]
############################################################################################################################
'''Simulate dataset for Bayesian Regression experiments'''
global n,nt,p,y,yt,X,Xt,tau
tau=args.tau
nw=args.n_data
p=args.data_dim
beta0=np.random.uniform(0.5,2,p)
idx0=set([item for item in range(p)])
S0=random.sample(idx0,int(p*(1-args.sparse)))
if p > 10:
    beta0[S0]=0.0
mean=np.zeros((p,))
rho=args.rho
cov=(1-rho)*np.identity(p)+rho*np.ones((p,p))
Xw= np.random.multivariate_normal(mean, cov, nw)
ew=np.random.normal(0,1,nw)
sigma = args.sigma
yw=np.dot(Xw,beta0)+sigma*ew
#print("yw",yw)
idx=set([item for item in range(nw)])
S=random.sample(idx,int(nw*0.8))
St=list(idx.difference(S))
X=torch.tensor(Xw[S,])
Xt=torch.tensor(Xw[St,])
y=torch.tensor(yw[S])
yt=torch.tensor(yw[St])
n=X.shape[0]
nt=Xt.shape[0]

y_y=torch.sum(y**2)
X_y=torch.sum(X.T*y,1)
X_X=torch.mm(X.T,X)
X_diag=torch.diag(X_X)

beta_ols = np.array(torch.matmul(torch.inverse(X_X),X_y))

#Function to calculate model rmse
def mod_rmse(y,x,beta):
    ypred = np.dot(x,beta.mean(0))
    return np.sqrt(np.mean((y - ypred)**2)) #np.sqrt(np.mean((y - ypred)**2))

set_all_seeds(args.seed)

rmse = {"Gibbs":None,"Flows":None,"MF-VI":None}
Time = {"Gibbs":None,"Flows":None,"MF-VI":None}

if args.sigma_prior == False:
    fig, axes = plt.subplots(1, 2, figsize=(6,3))
else:
    fig, axes = plt.subplots(1, 3, figsize=(7,2))


####################################################################GIBBS#############################################################################################
'''Gibbs Sampling'''
if args.mcmc=="gibbs":
    start = time.time()
    if args.sigma_prior == False:
        sigma_MCMC = 1
        Sig_MCMC=torch.inverse(X_X/(sigma_MCMC**2)+torch.eye(p)/tau**2) #inversion
        mu_MCMC=torch.sum(Sig_MCMC*X_y/(sigma_MCMC**2),1)
        beta_samples_MCMC_store=ss.multivariate_normal.rvs(mu_MCMC,Sig_MCMC,epochs_gibbs)
        idx = [i for i in range(0,beta_samples_MCMC_store.shape[0],thin)]
        beta_samples_MCMC_store=beta_samples_MCMC_store[idx]
    else:
        nsweep = epochs_gibbs
        sigma_samples_MCMC_store = np.zeros(nsweep)
        beta_samples_MCMC_store = np.zeros((nsweep,p))
        sigma_MCMC = ss.invgamma.rvs(a=2.5/2,scale=2.5/2)
        beta_MCMC = beta_ols
        for i in range(nsweep):
            #Update steps
            ytilda = y - np.dot(X,beta_MCMC)
            sigma_MCMCsq = ss.invgamma.rvs(a=(2.5/2+n/2),scale=(2.5/2+torch.sum(ytilda**2)/2))
            #if sigma_MCMCsq>5:
            #    print("IG params | sigmasq",(2.5/2+n/2),(2.5/2+torch.sum(ytilda**2)/2),sigma_MCMCsq)
            Sig_MCMC=torch.inverse(X_X/sigma_MCMCsq+torch.eye(p)/tau**2)
            mu_MCMC=torch.sum(Sig_MCMC*X_y/sigma_MCMCsq,1)
            beta_MCMC=ss.multivariate_normal.rvs(mu_MCMC,Sig_MCMC)
            #store values
            beta_samples_MCMC_store[i,] = beta_MCMC
            sigma_samples_MCMC_store[i] = sigma_MCMCsq
        
        burn_in=1000
        idx = [i for i in range(burn_in,beta_samples_MCMC_store.shape[0],thin)]
        beta_samples_MCMC_store=beta_samples_MCMC_store[idx]
        sigma_samples_MCMC_store=sigma_samples_MCMC_store[idx]
        sns.kdeplot(sigma_samples_MCMC_store,color='red',label='MCMC',ax=axes[2])
 
    Time["Gibbs"] = time.time()-start
    rmse["Gibbs"]= mod_rmse(np.array(yt),np.array(Xt),beta_samples_MCMC_store)

    #Plots
    
    sns.kdeplot(beta_samples_MCMC_store[:,0],color='red',label='MCMC',ax=axes[0])
    sns.kdeplot(beta_samples_MCMC_store[:,1],color='red',label='MCMC',ax=axes[1])
    
    rmse_betas_mcmc = np.sum((beta_samples_MCMC_store - beta0)**2,1)


    rmse_ypred_mcmc = np.sum((np.dot(Xt,beta_samples_MCMC_store.T)-np.expand_dims(yt.numpy(),1))**2,0)



########################################################################################################################################################
'''Some useful functions'''
X_y_f = X_y.type(torch.FloatTensor)
X_X_f = X_X.type(torch.FloatTensor)
y_y_f = y_y.type(torch.FloatTensor)



def gaussian_log_pdf(z):
    """
    Arguments:
    ----------
        - z: a batch of m data points (size: m x no. of params) tensor
    """
    return -.5 * (torch.log(torch.tensor([math.pi * 2], device=z.device)) + z ** 2).sum(1)

def invgamma_log_pdf(z,a,b):
    """
    Arguments:
    ----------
        - z: a batch of m data points (size: m x no. of params) tensor
    """
    a = torch.Tensor([a])
    b = torch.Tensor([b])
    return a*torch.log(b) - torch.lgamma(a) -(a+1)*torch.log(z) - b/z

def U5(z,s=None): #s is sigma_squared
    if s is None:
        return 0.5*(torch.sum(torch.mm(z, X_X_f)*z,1)-2*torch.sum(z*X_y_f,1)+y_y_f)
    else:
        return 0.5*n*torch.log(s) + 0.5*(torch.sum(torch.mm(z, X_X_f)*z,1)/s-2*torch.sum(z*X_y_f,1)/s+y_y_f/s)

def U6(z,s = None):
    if s is None:
        return 0.5*torch.sum(z**2,1)/tau**2
    else:
        return 0.5*torch.sum(z**2,1)/tau**2 - invgamma_log_pdf(s,2.5/2,2.5/2)

exact_log_density = lambda z,s=None: -U5(z,s)-U6(z,s)


class model(object):
    
    def __init__(self, target_energy, flow_type, ds_dim,cmade_dim):
#        self.mdl = nn_.SequentialFlow( 
#                flows.IAF(2, 128, 1, 3), 
#                flows.FlipFlow(1), 
#                flows.IAF(2, 128, 1, 3),
#                flows.FlipFlow(1), 
#                flows.IAF(2, 128, 1, 3))

        self.cmade_dim=cmade_dim
        if args.sigma_prior==False:
            self.nparams = p
        else: 
            self.nparams = p+1

        self.ds_dim = ds_dim

        if flow_type == "DSF":
            self.mdl = flows.IAF_DSF(self.nparams, self.cmade_dim, 1, 4,
                num_ds_dim=self.ds_dim, num_ds_layers=4)
        else:
            self.mdl = flows.IAF_DDSF(self.nparams, self.cmade_dim, 1, 4,
                num_ds_dim=self.ds_dim, num_ds_layers=4)

        
        self.optim = optim.Adam(self.mdl.parameters(), lr=args.lr, 
                                betas=(0.9, 0.999))
        
        self.target_energy = target_energy
        
    def train(self,epochs):
        
        total = epochs
        loss_store=[]
        for it in range(total):
            #print("epoch",it)

            self.optim.zero_grad()
            
            spl, logdet, _ = self.mdl.sample(64)
            if args.sigma_prior == False:
                losses = - self.target_energy(spl) - logdet 
            else:
                slices = list(range(0,p))
                losses = - self.target_energy(spl[:,slices],(torch.log(1+torch.exp(spl[:,-1])))**2) - logdet - torch.log(torch.exp(spl[:,-1])/(1+torch.exp(spl[:,-1]))) -torch.log(2*torch.log(1+torch.exp(spl[:,-1])))
            loss = losses.mean()
            
            loss.backward()
            self.optim.step()

            loss_store.append(loss.detach().numpy())


            
            if ((it + 1) % 1000) == 0:
                print ('Iteration: [%4d/%4d] loss: %.8f' % \
                    (it+1, total, loss.data))
        return loss_store

#######################################################MEAN-FIELD########################################################################################

y_y=torch.sum(y**2)
X_y=torch.sum(X.T*y,1)
X_X=torch.mm(X.T,X)
X_diag=torch.diag(X_X)

def elbo(params):
    tau_t = torch.Tensor([tau])
    beta_mu = params[0:p]
    beta_rho = params[(p):(2*p)]
    beta_sig = torch.log(1 + torch.exp(beta_rho))

    nll = 0
    nll = nll + 0.5*y_y
    nll = nll + 0.5*torch.sum(torch.sum(X*beta_mu,1)**2)
    nll = nll - torch.sum(beta_mu*X_y) 
    nll = nll + 0.5*torch.sum(X_diag*(beta_sig**2))
    kl_beta = 0.5*(torch.sum(beta_sig**2)/tau_t**2-2*torch.sum(torch.log(beta_sig)) + 2*p*torch.log(tau_t)+torch.sum(beta_mu**2)/tau_t**2-p)
    kl_sigma = 0
   
    if args.sigma_prior == True:
        sigma_mu = params[2*p]
        sigma_rho = params[2*p+1]
        sigma_rho_trans = torch.log(1 + torch.exp(sigma_rho))

        nll = nll*torch.exp(-sigma_mu + 0.5*sigma_rho_trans**2)
        nll = nll + 0.5*n*sigma_mu
        a = b = 2.5/2
        kl_sigma = a*sigma_mu - torch.log(sigma_rho_trans) + b*torch.exp(-sigma_mu + 0.5*sigma_rho_trans**2)
 
    kl = kl_sigma + kl_beta
    loss= nll+kl
    return loss

set_all_seeds(args.seed)

params1 = torch.Tensor(np.array(torch.matmul(torch.inverse(X_X),X_y)))
params2 = torch.Tensor(p).uniform_(-1,-1)
if args.sigma_prior == True:
    params3 = torch.Tensor(1).uniform_(-1,1)
    params4 = torch.Tensor(1).uniform_(-1,-1)
    params = torch.cat((params1,params2,params3,params4),0)
else:
    params = torch.cat((params1,params2),0)


start=time.time()
params.requires_grad_()
n_optim_steps = int(epochs_vi)
optimizer = torch.optim.Adam([params], args.lr)
start = time.time()
loss_store_vi=[]
for ii in range(n_optim_steps):
    optimizer.zero_grad()
    loss = elbo(params)
    loss_store_vi.append(loss.detach().numpy())
    print('Step # {}, loss: {}'.format(ii, loss.item()))
    loss.backward()
    # Access gradient if necessary
    optimizer.step()
Time["MF-VI"] = time.time()-start

mu_VI=np.array(params[0:p].detach())
Sig_VI=np.array((torch.eye(p)*(torch.log(1+torch.exp(params[p:(2*p)])))**2).detach())
beta_samples_VI=ss.multivariate_normal.rvs(mu_VI,Sig_VI,10000)
sns.kdeplot(beta_samples_VI[:,0],color='blue',label='MF-VI',ax=axes[0])
sns.kdeplot(beta_samples_VI[:,1],color='blue',label='MF-VI',ax=axes[1])
if args.sigma_prior==True:
    ln_sigma_samples_VI = np.log(1+np.exp(params[2*p+1].detach().numpy()))*np.random.normal(size=10000) + params[2*p].detach().numpy()
    sigma_samples_VI = np.exp(ln_sigma_samples_VI)
    sns.kdeplot(sigma_samples_VI,color='blue',label='MF-VI',ax=axes[2])

rmse["MF-VI"]= mod_rmse(np.array(yt),np.array(Xt),beta_samples_VI)
rmse_betas_vi = np.sum((beta_samples_VI - beta0)**2,1)
rmse_ypred_vi = np.sum((np.dot(Xt,beta_samples_VI.T)-np.expand_dims(yt.numpy(),1))**2,0)


################################################FLOWS###########################################################################
if args.flows=='NAF':
    set_all_seeds(args.seed)
    mdl = model(exact_log_density, "DSF", 8,cmade_dim)
    start = time.time()
    loss_store = mdl.train(epochs_flows)
    Time["Flows"]=time.time() - start
    data = mdl.mdl.sample(10000)[0].data.numpy()
    sns.kdeplot(data[:,0],color='green',label='Flows',ax=axes[0])
    sns.kdeplot(data[:,1],color='green',label='Flows',ax=axes[1])
    if args.sigma_prior==True:
        sns.kdeplot((np.log(1+np.exp(data[:,-1])))**2,color='green',label='Flows',ax=axes[2])
    rmse["Flows"]= mod_rmse(np.array(yt),np.array(Xt),data[:,list(range(0,p))])

if p==2: 
    plt.legend(bbox_to_anchor = (2.50, 0.6), loc='center right')   
    fig.tight_layout()
    fig.savefig(args.out+'beta_'+'trial'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()


rmse_betas_flows = np.sum((data[:,0:p] - beta0)**2,1)
rmse_ypred_flows = np.sum((np.dot(Xt,data[:,0:p].T)-np.expand_dims(yt.numpy(),1))**2,0)


#######################################PLOTS###########################################################################################################
if args.sigma_prior==True:
    plt.figure(figsize=(6.4,4.8))
    if args.mcmc=="gibbs":
        sns.kdeplot(sigma_samples_MCMC_store,color='red',label='MCMC')
    sns.kdeplot(sigma_samples_VI,color='blue',label='MF-VI')
    sns.kdeplot((np.log(1+np.exp(data[:,-1])))**2,color='green',label='Flows')
    plt.xlabel(r'$\sigma^{2}$')
    plt.legend()
    plt.tight_layout()
    plt.savefig(args.out+'sigmasq'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
    plt.clf()
    #print("sigmasq mcmc",np.mean(sigma_samples_MCMC_store))
    #print("sigmasq flows",np.mean((np.log(1+np.exp(data[:,-1])))**2))



if p==2:
    fig, ax = plt.subplots()
    handles =[]
    sns.kdeplot(data[:,0],data[:,1],color='green',label='Flows',ax=ax)
    handles.append(mlines.Line2D([], [], color='green', label="Flows"))
    sns.kdeplot(beta_samples_VI[:,0],beta_samples_VI[:,1],color='blue',label='MF-VI',ax=ax)
    handles.append(mlines.Line2D([], [], color='blue', label="MF-VI"))
    if args.mcmc=="gibbs":
        sns.kdeplot(beta_samples_MCMC_store[:,0],beta_samples_MCMC_store[:,1],color='red',label='Gibbs',ax=ax)
        handles.append(mlines.Line2D([], [], color='red', label="Gibbs"))
    ax.legend(handles = handles)
    plt.xlabel(r'$\beta_{1}$')
    plt.ylabel(r'$\beta_{2}$')
    plt.tight_layout()
    plt.savefig(args.out+'contour'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
    plt.clf()

if args.flows != 'None':
    plt.plot(loss_store)
    plt.savefig(args.out+'flows_loss_seed_'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
    plt.clf()

plt.plot(loss_store_vi)
plt.savefig(args.out+'vi_loss_seed_'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()


if args.mcmc=="gibbs":
    sns.kdeplot(rmse_betas_mcmc,color='red',label='MCMC')
sns.kdeplot(rmse_betas_flows,color='green',label='Flows')
sns.kdeplot(rmse_betas_vi,color='blue',label='MF-VI')
plt.legend()
plt.ylabel('Density of '+r'$||\beta-\beta_{0}||_{2}^{2}$')
plt.tight_layout()
plt.savefig(args.out+'rmse_beta_dist'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()

if args.mcmc=="gibbs":
    sns.kdeplot(rmse_ypred_mcmc,color='red',label='MCMC')
sns.kdeplot(rmse_ypred_flows,color='green',label='Flows')
sns.kdeplot(rmse_ypred_vi,color='blue',label='MF-VI')
plt.legend()
plt.ylabel('Density of '+r'$||\hat{y}_{\beta}-y_{true}||_{2}^{2}$')
plt.tight_layout()
plt.savefig(args.out+'rmse_ypred_dist'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()


print("n_data",args.n_data)
print("data_dim p",args.data_dim)
print("rmse",rmse)
print("Time in s",Time)

if args.mcmc!='None' and args.flows!='None':
    ypred_mcmc = np.dot(np.array(Xt),beta_samples_MCMC_store.mean(0))    
    ypred_flows = np.dot(np.array(Xt),data[:,list(range(0,p))].mean(0))
    ypred_vi = np.dot(np.array(Xt),beta_samples_VI.mean(0))
    fig, axes = plt.subplots(1, 3, figsize=(6,2))
    axes[0].scatter(ypred_mcmc,yt.numpy())
    axes[1].scatter(ypred_flows,yt.numpy())
    axes[2].scatter(ypred_vi,yt.numpy())
    #axes[0].set(xlabel="Fitted",ylabel="Actual")
    #axes[1].set(xlabel="Fitted",ylabel="Actual")
    #axes[2].set(xlabel="Fitted",ylabel="Actual")
    fig.supxlabel('Fitted')
    fig.supylabel('Actual')
   
    fig.tight_layout()
    plt.savefig(args.out+'actualvsfitted_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
    plt.clf()

##MCMC DIAGNOSTIC PLOTS###

if args.mcmc=="gibbs":

    if args.sigma_prior==True:
        fig = plt.figure(figsize=(6,2))       
        ax = fig.add_subplot(1,3,1)
        ax.plot(beta_samples_MCMC_store[:,0])
        ax = fig.add_subplot(1,3,2)
        ax.plot(beta_samples_MCMC_store[:,1])
        ax = fig.add_subplot(1,3,3)
        ax.plot(sigma_samples_MCMC_store)
        plt.savefig(args.out+'MH_Diagnostics/trace_seed_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
        plt.clf()
        
        fig, ax = plt.subplots(1,3) 
        sm.graphics.tsa.plot_acf(pd.DataFrame(beta_samples_MCMC_store[:,0]), lags=10, ax = ax[0])
        sm.graphics.tsa.plot_acf(pd.DataFrame(beta_samples_MCMC_store[:,1]), lags=10, ax = ax[1])
        sm.graphics.tsa.plot_acf(pd.DataFrame(sigma_samples_MCMC_store), lags=10, ax = ax[2])
    else:
        fig = plt.figure(figsize=(6,3))       
        ax = fig.add_subplot(1,2,1)
        ax.plot(beta_samples_MCMC_store[:,0])
        ax = fig.add_subplot(1,2,2)
        ax.plot(beta_samples_MCMC_store[:,1])
        plt.savefig(args.out+'MH_Diagnostics/trace_seed_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
        plt.clf()
    
        fig, ax = plt.subplots(1,2) 
        sm.graphics.tsa.plot_acf(pd.DataFrame(beta_samples_MCMC_store[:,0]), lags=10, ax = ax[0])
        sm.graphics.tsa.plot_acf(pd.DataFrame(beta_samples_MCMC_store[:,1]), lags=10, ax = ax[1])

    plt.savefig(args.out+"MH_Diagnostics/Autocorr_ndata_"+str(args.n_data)+'_p_'+str(args.data_dim)+".png")
    plt.clf()


#####################################################INFERENCE##########################################################################
if args.mcmc=="gibbs":
    q_lower_mcmc = np.quantile(beta_samples_MCMC_store,0.025,axis=0)
    q_upper_mcmc = np.quantile(beta_samples_MCMC_store,0.975,axis=0)
q_lower_vi = np.quantile(beta_samples_VI,0.025,axis=0)
q_upper_vi = np.quantile(beta_samples_VI,0.975,axis=0)
q_lower_flows = np.quantile(data[:,0:p],0.025,axis=0)
q_upper_flows = np.quantile(data[:,0:p],0.975,axis=0)

#Checking if true beta0 lies in estimated CI
correct_pct = {"Gibbs":None,"Flows":None,"MF-VI":None}
if args.mcmc=="gibbs":
    cint_mcmc = (q_lower_mcmc<= beta0) & (beta0 <= q_upper_mcmc) 
    correct_pct["Gibbs"] = sum(cint_mcmc)/(cint_mcmc).shape[0]

cint_vi = (q_lower_vi<= beta0) & (beta0 <= q_upper_vi)
cint_flows = (q_lower_flows<= beta0) & (beta0 <= q_upper_flows)
correct_pct["MF-VI"] = sum(cint_vi)/(cint_vi).shape[0]
correct_pct["Flows"]= sum(cint_flows)/(cint_flows).shape[0]

#Testing null hypothesis beta0 = 0
confusion = {"Gibbs":None,"Flows":None,"MF-VI":None}
ht_beta0 = beta0!=0
if args.mcmc=="gibbs":
    ht_mcmc = (q_lower_mcmc> np.zeros(p)) | (np.zeros(p) > q_upper_mcmc) 
    confusion["Gibbs"] = metrics.confusion_matrix(ht_beta0,ht_mcmc)

ht_vi = (q_lower_vi > np.zeros(p)) | (np.zeros(p) > q_upper_vi)
ht_flows = (q_lower_flows > np.zeros(p)) | (np.zeros(p) > q_upper_flows)
confusion["MF-VI"] = metrics.confusion_matrix(ht_beta0,ht_vi)
confusion["Flows"] = metrics.confusion_matrix(ht_beta0,ht_flows)

print("Correct prop",correct_pct)
print("Confusion Gibbs",confusion["Gibbs"])
print("Confusion MF-VI",confusion["MF-VI"])
print("Confusion Flows",confusion["Flows"])


#############Write Results to Excel######################################################

if args.writecsv==True:
    grid_to_row = {'2_50':0,'2_100':1, '2_200':2, '20_50':3,'20_100':4,'20_200':5,
    '50_50':6,'50_100':7,'50_200':8,'100_50':9,'100_100':10,'100_200':11}
    key = grid_to_row[str(args.data_dim)+'_'+str(args.n_data)]
    from openpyxl import workbook 
    from openpyxl import load_workbook
    filepath = args.out+"Result.xlsx"
    wb = load_workbook(filepath)
    sheets = wb.sheetnames
    main = wb['Main']
    main.cell(row = key+2, column = 2).value = Time["Gibbs"]
    main.cell(row = key+2, column = 3).value = Time["Flows"]
    main.cell(row = key+2, column = 4).value = Time["MF-VI"]
    main.cell(row = key+2, column = 5).value = rmse["Gibbs"]
    main.cell(row = key+2, column = 6).value = rmse["Flows"]
    main.cell(row = key+2, column = 7).value = rmse["MF-VI"]
    main.cell(row = key+2, column = 8).value = correct_pct["Gibbs"]
    main.cell(row = key+2, column = 9).value = correct_pct["Flows"]
    main.cell(row = key+2, column = 10).value = correct_pct["MF-VI"]
    confusion_gibbs=wb['Confusion_Gibbs']
    confusion_flows=wb['Confusion_Flows']
    confusion_vi=wb['Confusion_VI']
    if confusion["Gibbs"] is not None:
        confusion_gibbs.cell(row=key+2,column=2).value=confusion["Gibbs"][0][0]
        if len(confusion["Gibbs"][0])>1:
            confusion_gibbs.cell(row=key+2,column=3).value=confusion["Gibbs"][0][1]
        if len(confusion["Gibbs"])>1:
            confusion_gibbs.cell(row=key+2,column=4).value=confusion["Gibbs"][1][0]
        if len(confusion["Gibbs"])>1 and len(confusion["Gibbs"][0])>1:
            confusion_gibbs.cell(row=key+2,column=5).value=confusion["Gibbs"][1][1]
    if confusion["Flows"] is not None:
        confusion_flows.cell(row=key+2,column=2).value=confusion["Flows"][0][0]
        if len(confusion["Flows"][0])>1:
            confusion_flows.cell(row=key+2,column=3).value=confusion["Flows"][0][1]
        if len(confusion["Flows"])>1:
            confusion_flows.cell(row=key+2,column=4).value=confusion["Flows"][1][0]
        if len(confusion["Flows"])>1 and len(confusion["Flows"][0])>1:
            confusion_flows.cell(row=key+2,column=5).value=confusion["Flows"][1][1]
    if confusion["MF-VI"] is not None:
        confusion_vi.cell(row=key+2,column=2).value=confusion["MF-VI"][0][0]
        if len(confusion["MF-VI"][0])>1:
            confusion_vi.cell(row=key+2,column=3).value=confusion["MF-VI"][0][1]
        if len(confusion["MF-VI"])>1:
            confusion_vi.cell(row=key+2,column=4).value=confusion["MF-VI"][1][0]
        if len(confusion["MF-VI"])>1 and len(confusion["MF-VI"][0])>1:
            confusion_vi.cell(row=key+2,column=5).value=confusion["MF-VI"][1][1]
    wb.save(filepath) 

