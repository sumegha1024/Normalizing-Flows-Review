
'''Generate energy density samples based on toy energy functions using flows and mh algorithms'''

import numpy as np
import torch
import random
import torch.optim as optim
from torch.autograd import Variable
import sys
import os
sys.path.append(os.path.join(sys.path[0],'Flows_scripts'))
import nn as nn_
import flows
import matplotlib.pyplot as plt
import time
import seaborn as sns
import random
import argparse
from utils import set_all_seeds, kldiv
import pandas as pd
import statsmodels.api as sm
from scipy import stats
sys.path.append(os.path.join(sys.path[0],'MH_Samplers'))
from MH_sampler_basic import mh
from toy_energy import U1, U2, U3, U4, U5, U6, U7, U8, U9

parser = argparse.ArgumentParser(description='inputs_toyenergy')
parser.add_argument('--flows', default = 'NAF', type=str, help='type of flow model; Use None if result for flows is not required')
parser.add_argument('--mcmc', default = 'mh', type=str, help='type of mcmc method; Use None if result for mcmc is not required')
parser.add_argument('--kdeplot', default=False, action='store_true', help='Whether we want kernel density plots or histograms for fitted model samples. Default corresponds to hist')
parser.add_argument('--heatmap', default=False, action='store_true', help='Whether we want heatmaps for kde error')
parser.add_argument('--key',default = 'U1', type=str, help='Which energy function to plot')
parser.add_argument('--seed',default =0, type=int)
parser.add_argument('--out',default = '/mnt/home/premchan/Normalizing-Flows-Review/Out/Out_toy/', type=str, help='path to results')
parser.add_argument('--writecsv', default=False, action='store_true', help='Whether we want to write metrics to Result.xlsx')

args = parser.parse_args()

set_all_seeds(args.seed)
#Dictionary of parameters used for each model
#Left to right: Energy function, number of epochs for flows, nsamps for MH, hidden layer dimension, Type of NAF flow, approximated normalization constant

toy_energy={"U1":[U1,15000,400000,5, "DSF",6.5371],
            "U2":[U2,15000,400000,5, "DSF",27.3805],
            "U3":[U3,15000,400000,2,"DSF", 31.5069],
            "U4":[U4,15000,400000,5,"DSF", 3.0456],
            "U5":[U5,15000,400000,5, "DSF",5.6198],
            "U6":[U6,15000,400000,5,"DSF",13.9084],
            "U7":[U7,15000,400000,8,"DSF",9.8347],
            "U8":[U8,15000,400000,8,"DSF",11.8317],
            "U9":[U9,15000,400000,16,"DSF",30.7431]
            }

class model(object):
    
    def __init__(self, target_energy, flow_type, ds_dim):
#        self.mdl = nn_.SequentialFlow( 
#                flows.IAF(2, 128, 1, 3), 
#                flows.FlipFlow(1), 
#                flows.IAF(2, 128, 1, 3),
#                flows.FlipFlow(1), 
#                flows.IAF(2, 128, 1, 3))

        self.ds_dim = ds_dim

        if flow_type == "DSF":
            self.mdl = flows.IAF_DSF(2, 64, 1, 4,
                num_ds_dim=self.ds_dim, num_ds_layers=2)
        else:
            self.mdl = flows.IAF_DDSF(2, 64, 1, 4,
                num_ds_dim=self.ds_dim, num_ds_layers=2)

        
        self.optim = optim.Adam(self.mdl.parameters(), lr=0.0005, 
                                betas=(0.9, 0.999))
        
        self.target_energy = target_energy
        
    def train(self,epochs):
        
        total = epochs
        loss_store=[]
        for it in range(total):

            self.optim.zero_grad()
            
            spl, logdet, _ = self.mdl.sample(64)
            losses = - self.target_energy(spl) - logdet
            loss = losses.mean()
            
            loss.backward()
            self.optim.step()

            loss_store.append(loss.detach().numpy())


            
            if ((it + 1) % 1000) == 0:
                print ('Iteration: [%4d/%4d] loss: %.8f' % \
                    (it+1, total, loss.data))
        return loss_store
             


# build and train
time_store={"MH":None,"Flows":None}
accept_store = []
error_kde = {"MH":None,"Flows":None}
error_kl1 = {"MH":None,"Flows":None}
error_kl2 = {"MH":None,"Flows":None}
error_kl_avg = {"MH":None,"Flows":None}
key = args.key
    
ef = toy_energy[key][0]
epochs = toy_energy[key][1]
nsamps = toy_energy[key][2]
ds_dim = toy_energy[key][3]
flow_type = toy_energy[key][4]
    
n=200
fig = plt.figure()       
ax = fig.add_subplot(1,3,1)
x = np.linspace(-10,10,n)
y = np.linspace(-10,10,n)
xx,yy = np.meshgrid(x,y)
X = np.concatenate((xx.reshape(n**2,1),yy.reshape(n**2,1)),1)
X = X.astype('float32')
X1 = Variable(torch.from_numpy(X))
Z = np.exp(ef(X1).data.numpy()).reshape(n,n)/toy_energy[key][5]
#print("Zshape",Z.shape)
   
ax.pcolormesh(xx,yy,np.exp(Z))
ax.axis('off')
plt.xlim((-10,10))
plt.ylim((-10,10))

'''MCMC'''
if args.mcmc == "mh":
        
    start = time.time()
    mh_samples,accept = mh(nsamps, 2, ef, burn_in=0)
    idx = [i for i in range(0,mh_samples.shape[0],40)]
    time_store["MH"]=(time.time() - start)
    mh_samples = mh_samples[idx,:].numpy()
    accept_store.append(accept)

    #Kernel density estimates
    kernel = stats.gaussian_kde(mh_samples.transpose())
    density = kernel(X.transpose()).reshape(xx.shape)
    kde_MH = np.abs(density - Z)
    error_kde["MH"] = np.sqrt(np.sum((density - Z)**2))
    error_kl1["MH"] = kldiv(density,Z)
    error_kl2["MH"] = kldiv(Z,density)
    error_kl_avg["MH"] = (error_kl1["MH"] + error_kl2["MH"])/2
        
    #Visualization
    XX = mh_samples[:,0]
    YY = mh_samples[:,1]
    ax = fig.add_subplot(1,3,2)
    if args.kdeplot == True:
        sns.kdeplot(XX,YY,cmap="mako",fill=True,ax=ax)
    else:
        plot = ax.hist2d(XX,YY,200,range=np.array([(-10, 10), (-10, 10)]),density=True)
        
        plt.xlim((-10,10))
        plt.ylim((-10,10))
        plt.axis('off')


'''flows'''
if args.flows != 'None':
    mdl = model(ef, flow_type, ds_dim)
    start = time.time()
    loss_store = mdl.train(epochs)
    time_store["Flows"] = time.time() - start

    if args.mcmc!='None':
        data = mdl.mdl.sample(len(idx))[0].data.numpy()
    else:
        data = mdl.mdl.sample(10000)[0].data.numpy()
       
    #Kernel density estimates
    kernel = stats.gaussian_kde(data.transpose())
    density = kernel(X.transpose()).reshape(xx.shape)
    kde_flows = np.abs(density - Z)
    error_kde["Flows"] = np.sqrt(np.sum((density - Z)**2))
    error_kl1["Flows"] = kldiv(density,Z)
    error_kl2["Flows"] = kldiv(Z,density)
    error_kl_avg["Flows"] = (error_kl1["Flows"] + error_kl2["Flows"])/2

       
    #Plot flows 
    ax = fig.add_subplot(1,3,3)
    XX = data[:,0]
    YY = data[:,1]
    if args.kdeplot == True:
        sns.kdeplot(XX,YY,cmap="mako",fill=True,ax=ax)
    else:
        plot = ax.hist2d(XX,YY,200,range=np.array([(-10, 10), (-10, 10)]),density=True)
        plt.xlim((-10,10))
        plt.ylim((-10,10))
        plt.axis('off')
    
plt.savefig(args.out+'model_toy'+ key + '_seed'+str(args.seed)+'.png')
plt.clf()

if args.mcmc != 'None':
    fig = plt.figure()       
    ax = fig.add_subplot(1,2,1)
    ax.plot(mh_samples[:,0])
    ax = fig.add_subplot(1,2,2)
    ax.plot(mh_samples[:,1])
    plt.savefig(args.out+'MH_Diagnostics/trace'+key+'_seed'+str(args.seed)+'.png')
    plt.clf()
    mh_samples = pd.DataFrame(mh_samples)
    fig, ax = plt.subplots(1,2)       
    sm.graphics.tsa.plot_acf(mh_samples.iloc[:,0], lags=50, ax = ax[0])
    sm.graphics.tsa.plot_acf(mh_samples.iloc[:,1], lags=50, ax = ax[1])
    plt.savefig(args.out+'MH_Diagnostics/auto_corr'+ key +'_seed'+str(args.seed)+'.png')
    plt.clf()
    
if args.flows != 'None':
    plt.plot(loss_store)
    plt.savefig(args.out+'loss_toy'+key+'_seed'+str(args.seed)+'.png')
    plt.clf()
    print("Seed",args.seed)
    print("Mean loss",np.array(loss_store[14000:15000]).mean())
    print("SD of loss",np.array(loss_store[14000:15000]).std())
    #loss_store_df = pd.DataFrame(loss_store)
    #loss_store_df.to_csv(args.out+'loss_'+ key +'_seed'+str(args.seed)+'.csv')
    #idx = [i for i in range(0,90000,40)]

if args.heatmap == True:
    fig, (ax1, ax2) = plt.subplots(ncols=2)
    fig.subplots_adjust(wspace=0.4)
    if args.mcmc != 'None':          
        #ax = fig.add_subplot(1,2,1)
        ss = np.arange(n)[::-1]
        sns.heatmap(kde_MH, cmap='RdBu_r', vmin=0, vmax=np.amax(kde_MH),ax = ax1)

    if args.flows != 'None':
        #ax = fig.add_subplot(1,2,2)
        ss = np.arange(n)[::-1]
        sns.heatmap(kde_flows, cmap='RdBu_r', vmin=0, vmax=np.amax(kde_flows), ax = ax2)        
    plt.savefig(args.out+"heatmap"+key+'_seed'+str(args.seed)+".png")

    plt.clf()

print(key)
print("Error_KDE",error_kde) 
print("improvement percent",(error_kde["MH"]-error_kde["Flows"])*100/error_kde["MH"])     
print("Time",time_store)
print("MH: Acceptance rates",accept_store)
print("Error_KL q,p",error_kl1)
print("Error_KL p,q",error_kl2)
print("Error_KL avg",error_kl_avg)


#############Write Results to Excel############################
if args.writecsv==True:
    from openpyxl import workbook 
    from openpyxl import load_workbook
    filepath = args.out+"Result.xlsx"
    wb = load_workbook(filepath)
    sheets = wb.sheetnames
    print(sheets)
    time_flows = wb[sheets[0]]
    time_flows.cell(row = int(list(args.key)[1])+1, column = args.seed+2).value = time_store["Flows"]
    time_mh = wb[sheets[1]]
    time_mh.cell(row = int(list(args.key)[1])+1, column = args.seed+2).value = time_store["MH"]
    kde_flows = wb[sheets[2]]
    kde_flows.cell(row = int(list(args.key)[1])+1, column = args.seed+2).value = error_kde["Flows"]
    kde_mh = wb[sheets[3]]
    kde_mh.cell(row = int(list(args.key)[1])+1, column = args.seed+2).value = error_kde["MH"]
    loss = wb[sheets[4]]
    loss.cell(row = int(list(args.key)[1])+1, column = args.seed+2).value = np.array(loss_store[14000:15000]).mean()
    wb.save(filepath) 


