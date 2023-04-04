'''Logistic Regression'''

import torch
import torch.optim as optim
import sys
import os
sys.path.append(os.path.join(sys.path[0],'Flows_scripts'))
import nn as nn_
import torch.nn as nn
from tqdm import tqdm
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import math
import scipy.stats as ss
from scipy import special
import seaborn as sns
import random
import time
from torch.distributions import MultivariateNormal
import argparse
import pandas as pd
import flows
from utils import set_all_seeds
from sklearn import metrics
import statsmodels.api as sm
sys.path.append(os.path.join(sys.path[0],'MH_Samplers'))
from MH_sampler_logistic import mh
from sklearn.linear_model import LogisticRegression


parser = argparse.ArgumentParser(description='inputs_logits')
parser.add_argument('--rho',default =0.0, type=float, help='Correlation coefficient for design matrix')
parser.add_argument('--tau',default =1.0, type=float, help='Standard dev for simulated data')
parser.add_argument('--lr', default = 0.0005, type=float, help='learning rate')
parser.add_argument('--seed',default =3, type=int, help='seed for simulation')
parser.add_argument('--out',default = '/mnt/home/premchan/Normalizing-Flows-Review/Temp/', type=str, help='path to results')
parser.add_argument('--data_dim',default =2, type=int, help='dimension of beta vector')
parser.add_argument('--n_data',default =50, type=int, help='number of samples for y')
parser.add_argument('--sparse',default =0.2, type=float, help='sparsity level in %')
parser.add_argument('--writecsv', default=False, action='store_true', help='Whether we want to write metrics to Result.xlsx')
#parser.add_argument('--flows_epochs',default =15000, type=int)

args = parser.parse_args()

print("n_data",args.n_data)
print("data_dim p",args.data_dim)
print("seed",args.seed)
set_all_seeds(0)


#tuning_dict= {'2_50':[200000,15000,1,64,20],'2_100':[200000,15000,0.8,64,20], '2_200':[200000,15000,0.5,64,20], '20_50':[1000000,15000,0.2,64,100],'20_100':[1000000,15000,0.15,64,100],'20_200':[1000000,15000,0.1,64,100],
#    '50_50':[2000000,50000,0.2,64,200],'50_100':[4000000,50000,0.15,64,400],'50_200':[4000000,50000,0.1,64,400],'100_50':[4000000,50000,0.1,128,400],'100_100':[5000000,50000,0.1,128,500],'100_200':[5000000,50000,0.08,128,500]}

tuning_dict= {'2_50':[200000,15000,1,64,20],'2_100':[200000,15000,0.8,64,20], '2_200':[200000,15000,0.5,64,20], '20_50':[1000000,15000,0.2,64,100],'20_100':[1000000,15000,0.15,64,100],'20_200':[1000000,15000,0.1,64,100],
    '50_50':[2000000,15000,0.2,64,200],'50_100':[4000000,15000,0.15,64,400],'50_200':[4000000,15000,0.1,64,400],'100_50':[4000000,15000,0.1,128,400],'100_100':[5000000,15000,0.1,128,500],'100_200':[5000000,15000,0.08,128,500]}


epochs_mh=tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][0]
epochs_flows =tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][1]
#epochs_flows=args.flows_epochs
mh_noise =tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][2]
cmade_dim=tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][3]
thin = tuning_dict[str(args.data_dim)+'_'+str(args.n_data)][4]

#beta_plot_pairs=[(0,1),(5,6),(27,19),(46,3),(17,12),(23,2),(6,35),(9,11),(5,21),(13,40)]
############################################################################################################################
'''Simulate dataset for Bayesian Regression experiments'''
global n,nt,p,y,X,Xt,tau
tau=args.tau
nw=args.n_data
p=args.data_dim
beta0=np.random.uniform(0.5,2.0,p)
idx0=set([item for item in range(p)])
S0=random.sample(idx0,int(p*(1-args.sparse)))
if p > 10:
    beta0[S0]=0.0
mean=np.zeros((p,))
rho=args.rho
cov=(1-rho)*np.identity(p)+rho*np.ones((p,p))
Xw= np.random.multivariate_normal(mean, cov, nw)
logitsw=np.dot(Xw,beta0)
pw = np.exp(logitsw)/(1+np.exp(logitsw))
yw=np.random.binomial(n=1,p=pw)
#print("yw",yw)
idx=set([item for item in range(nw)])
S=random.sample(idx,int(nw*0.8))
St=list(idx.difference(S))
X=torch.tensor(Xw[S,])
Xt=torch.tensor(Xw[St,])
y=torch.tensor(yw[S])
yt=torch.tensor(yw[St])
n=X.shape[0]
nt=Xt.shape[0]

lr_fit = LogisticRegression(penalty='none',fit_intercept=False,solver='sag').fit(X,y)
#print("fit",lr_fit.coef_.shape)

Time = {"MH":None,"Flows":None,"MF-VI":None}
acc = {"MH":None,"Flows":None,"MF-VI":None}
se_beta_avg = {"MH":None,"Flows":None,"MF-VI":None}
se_beta_min = {"MH":None,"Flows":None,"MF-VI":None}
se_beta_max = {"MH":None,"Flows":None,"MF-VI":None}


set_all_seeds(args.seed)
'''Some useful functions'''

sigmoid = nn.Sigmoid()

def mod_acc(y,x,beta):
    y=y.numpy()
    x=x.numpy()
    beta=beta.numpy()
    logitspred = np.dot(x,beta.mean(0))
    probits=np.exp(logitspred)/(1+np.exp(logitspred))
    out_labels=(probits>=0.5)
    ylabels=(y==1)
    count=(out_labels==ylabels)

    return sum(count)/len(count)

#NLL
def U5(z): 
    logits_mean= torch.matmul(X.type(torch.float32),torch.transpose(z,0,1))
    return -(y.unsqueeze(1)*logits_mean).sum(0) + torch.log1p(torch.exp(logits_mean)).sum(0)

#Negative Log prior for beta
def U6(z):
    return 0.5*torch.sum(z**2,1)/tau**2

'''Input z should be torch tensor dimension nsamps x data_dim'''
exact_log_density = lambda z: -U5(z)-U6(z) #Joint log density for flows

#########################################################MCMC###################################################################
'''Random Walk Metropolis Hastings'''
fig, axes = plt.subplots(1, 2, figsize=(6,3))
start = time.time()
beta_samples_mh_store,accept = mh(epochs_mh, p, exact_log_density,burn_in=0,sigma=mh_noise,log_trans=False,log_scale=True,current_init=lr_fit.coef_)
Time["MH"] = time.time()-start
idx = [i for i in range(0,beta_samples_mh_store.size()[0],thin)]
beta_samples_mh_store=beta_samples_mh_store[idx,:]
sns.kdeplot(beta_samples_mh_store[:,0].numpy(),color='red',label='MCMC',ax=axes[0])
sns.kdeplot(beta_samples_mh_store[:,1].numpy(),color='red',label='MCMC',ax=axes[1])

se_beta_avg["MH"] =np.mean(np.std(beta_samples_mh_store.numpy(),axis=0,ddof=1))
se_beta_min["MH"] =np.min(np.std(beta_samples_mh_store.numpy(),axis=0,ddof=1))
se_beta_max["MH"] =np.max(np.std(beta_samples_mh_store.numpy(),axis=0,ddof=1))

sse_betas_mcmc = np.sum((beta_samples_mh_store.numpy() - beta0)**2,1)
acc["MH"]=mod_acc(yt,Xt,beta_samples_mh_store)

logitspred_mcmc = np.dot(Xt,beta_samples_mh_store.T)
probits_mcmc=np.exp(logitspred_mcmc)/(1+np.exp(logitspred_mcmc))
out_labels_mcmc=(probits_mcmc>=0.5)
ytlabels=(yt.numpy()==1)
count_mcmc=(out_labels_mcmc.T==ytlabels)
acc_samples_mcmc =count_mcmc.sum(1)/len(ytlabels)

#######################################################MEAN-FIELD####################################################################

class logisticreg(nn.Module):
    def __init__(self,data_dim,beta_init):
        super().__init__()
        
        self.beta_init=beta_init
        self.dim = data_dim
        self.beta_mu=nn.Parameter(self.beta_init.squeeze())
        self.beta_rho=nn.Parameter(torch.Tensor(self.dim).uniform_(-1,-1))
        self.betas = torch.zeros(self.dim)
        self.kl=0

    def forward(self, x):
        beta_sig = torch.log(1 + torch.exp(self.beta_rho))
        self.betas = beta_sig*(torch.Tensor(self.dim).normal_(0,1))+self.beta_mu 
        

        #Define kl 
        tau_t = torch.Tensor([tau])
        kl_beta = 0.5*(torch.sum(beta_sig**2)/tau_t**2-2*torch.sum(torch.log(beta_sig)) + 2*p*torch.log(tau_t)+torch.sum(self.beta_mu**2)/tau_t**2-p)
        self.kl=kl_beta
        #Return logits
        logits=torch.sum((x*self.betas),1)

        return logits
    
    def elbo(self,y,x,n_samples):
        n=len(y)
        logits = torch.zeros(n_samples,n)
        loglik = torch.zeros(n_samples)
        kll = torch.zeros(n_samples)
        for i in range(n_samples):
            logits[i] = self.forward(x)
            kll[i] = self.kl
            loglik[i]=torch.sum(((y*logits[i]) - torch.log1p(torch.exp(logits[i]))),0) 
            
        return sum(kll)/n_samples-sum(loglik)/(n_samples), logits
        
set_all_seeds(args.seed)
                
mod = logisticreg(args.data_dim,torch.Tensor(lr_fit.coef_))
n_optim_steps = int(30000)
optimizer = torch.optim.Adam(mod.parameters(), args.lr)
start = time.time()
loss_store_vi=[]

for ii in range(n_optim_steps):
    optimizer.zero_grad()
    loss,_ = mod.elbo(y,X,8)
    loss_store_vi.append(loss.detach().numpy())
    print('Step # {}, loss: {}'.format(ii, loss.item()))
    loss.backward()
    optimizer.step()
Time["MF-VI"] = time.time()-start

beta_samples_VI = torch.zeros(10000,p)
for i in range(10000):
    mod.forward(Xt)
    beta_samples_VI[i,:]=mod.betas

acc["MF-VI"]=mod_acc(yt,Xt,beta_samples_VI.detach())
beta_samples_VI=beta_samples_VI.detach().numpy()
sns.kdeplot(beta_samples_VI[:,0],color='blue',label='MF-VI',ax=axes[0])
sns.kdeplot(beta_samples_VI[:,1],color='blue',label='MF-VI',ax=axes[1])

se_beta_avg["MF-VI"] =np.mean(np.std(beta_samples_VI,axis=0,ddof=1))
se_beta_min["MF-VI"] =np.min(np.std(beta_samples_VI,axis=0,ddof=1))
se_beta_max["MF-VI"] =np.max(np.std(beta_samples_VI,axis=0,ddof=1))

sse_betas_vi = np.sum((beta_samples_VI - beta0)**2,1)
logitspred_vi = np.dot(Xt,beta_samples_VI.T)
probits_vi=np.exp(logitspred_vi)/(1+np.exp(logitspred_vi))
out_labels_vi=(probits_vi>=0.5)
ytlabels=(yt.numpy()==1)
count_vi=(out_labels_vi.T==ytlabels)
acc_samples_vi =count_vi.sum(1)/len(ytlabels)



####################################Flows################################################################################

class model(object):
    
    def __init__(self, target_energy, flow_type, ds_dim,cmade_dim):
#        self.mdl = nn_.SequentialFlow( 
#                flows.IAF(2, 128, 1, 3), 
#                flows.FlipFlow(1), 
#                flows.IAF(2, 128, 1, 3),
#                flows.FlipFlow(1), 
#                flows.IAF(2, 128, 1, 3))

        self.nparams = p
        self.ds_dim = ds_dim
        self.cmade_dim=cmade_dim

        if flow_type == "DSF":
            self.mdl = flows.IAF_DSF(self.nparams, self.cmade_dim, 1, 4,
                num_ds_dim=self.ds_dim, num_ds_layers=4)
        else:
            self.mdl = flows.IAF_DDSF(self.nparams, self.cmade_dim, 1, 4,
                num_ds_dim=self.ds_dim, num_ds_layers=4)

        
        self.optim = optim.Adam(self.mdl.parameters(), lr=args.lr, 
                                betas=(0.9, 0.999))
        
        self.target_energy = target_energy
        
    def train(self,epochs):
        
        total = epochs
        loss_store=[]
        for it in range(total):
            #print("epoch",it)

            self.optim.zero_grad()
            
            spl, logdet, _ = self.mdl.sample(64)
            
            losses = - self.target_energy(spl) - logdet 
            loss = losses.mean()
            
            loss.backward()
            self.optim.step()

            loss_store.append(loss.detach().numpy())


            
            if ((it + 1) % 1000) == 0:
                print ('Iteration: [%4d/%4d] loss: %.8f' % \
                    (it+1, total, loss.data))
        return loss_store


set_all_seeds(args.seed)
mdl = model(exact_log_density, "DSF", 8,cmade_dim)
start = time.time()
loss_store = mdl.train(epochs_flows)
Time["Flows"]=time.time() - start
data = mdl.mdl.sample(10000)[0].data.numpy()
sns.kdeplot(data[:,0],color='green',label='Flows',ax=axes[0])
if p>2: axes[0].axvline(x=0, color='purple',ls='--')
axes[0].axvline(x=beta0[0], color='gray',ls='--', label='Truth')
sns.kdeplot(data[:,1],color='green',label='Flows',ax=axes[1])
if p>2: axes[1].axvline(x=0, color='purple',ls='--')
axes[1].axvline(x=beta0[1], color='gray',ls='--', label='Truth')
acc["Flows"]=mod_acc(yt,Xt,beta_samples_mh_store)
se_beta_avg["Flows"] =np.mean(np.std(data[:,0:p],axis=0,ddof=1))
se_beta_min["Flows"] =np.min(np.std(data[:,0:p],axis=0,ddof=1))
se_beta_max["Flows"] =np.max(np.std(data[:,0:p],axis=0,ddof=1))

logitspred_flows = np.dot(Xt,data.T)
probits_flows=np.exp(logitspred_flows)/(1+np.exp(logitspred_flows))
out_labels_flows=(probits_flows>=0.5)
count_flows=(out_labels_flows.T==ytlabels)
acc_samples_flows =count_flows.sum(1)/len(ytlabels)
sse_betas_flows = np.sum((data[:,0:p] - beta0)**2,1)


################################################PLOTS##############################################################

plt.legend(bbox_to_anchor = (1.75, 0.6), loc='center right')   
fig.tight_layout()
fig.savefig(args.out+'beta_'+'trial'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()

fig, ax = plt.subplots()
handles =[]
sns.kdeplot(data[:,0],data[:,1],color='green',label='Flows',ax=ax)
handles.append(mlines.Line2D([], [], color='green', label="Flows"))
sns.kdeplot(beta_samples_VI[:,0],beta_samples_VI[:,1],color='blue',label='MF-VI',ax=ax)
handles.append(mlines.Line2D([], [], color='blue', label="MF-VI"))
sns.kdeplot(beta_samples_mh_store[:,0],beta_samples_mh_store[:,1],color='red',label='MCMC',ax=ax)
handles.append(mlines.Line2D([], [], color='red', label="MCMC"))
ax.legend(handles = handles,fontsize=15)
plt.xlabel(r'$\beta_{1}$',fontsize=20)
plt.ylabel(r'$\beta_{2}$',fontsize=20)
plt.yticks(fontsize=10)
plt.xticks(fontsize=10)
plt.tight_layout()
plt.savefig(args.out+'contour'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.pdf')
plt.clf() 


##PAIRWISE PLOTS OF BETA PARAMS
beta_pairs=False
if beta_pairs==True:
    for tuples in beta_plot_pairs:
        fig, axes = plt.subplots(1, 2, figsize=(6,3))   
        sns.kdeplot(beta_samples_mh_store[:,tuples[0]].numpy(),color='red',label='MCMC',ax=axes[0])
        sns.kdeplot(beta_samples_mh_store[:,tuples[1]].numpy(),color='red',label='MCMC',ax=axes[1])
        sns.kdeplot(beta_samples_VI[:,tuples[0]],color='blue',label='MF-VI',ax=axes[0])
        sns.kdeplot(beta_samples_VI[:,tuples[1]],color='blue',label='MF-VI',ax=axes[1])
        sns.kdeplot(data[:,tuples[0]],color='green',label='Flows',ax=axes[0])
        sns.kdeplot(data[:,tuples[1]],color='green',label='Flows',ax=axes[1])
        fig.tight_layout()
        fig.savefig(args.out+'beta_pair'+str(tuples[0])+'_'+str(tuples[1])+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
        plt.legend()
    plt.clf()

    for tuples in beta_plot_pairs:
        sns.kdeplot(data[:,tuples[0]],data[:,tuples[1]],color='green',label='Flows')
        sns.kdeplot(beta_samples_mh_store[:,tuples[0]],beta_samples_mh_store[:,tuples[1]],color='red',label='MCMC')
        sns.kdeplot(beta_samples_VI[:,tuples[0]],beta_samples_VI[:,tuples[1]],color='blue',label='MF-VI') 
        plt.legend()
        plt.xlabel(r'$\beta_{1}$')
        plt.ylabel(r'$\beta_{2}$')
        plt.savefig(args.out+'contour_beta_pair'+str(tuples[0])+'_'+str(tuples[1])+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
        plt.clf()

##LOSS PLOTS
#loss_store_df=pd.DataFrame(loss_store)
#loss_store_df.to_csv(args.out+"loss_flows_ndata_"+str(args.n_data)+"_p_"+str(args.data_dim)+".csv")
#print("Flows_epochs, Loss mean",epochs_flows,np.mean(loss_store[(epochs_flows-5000):epochs_flows]))

plt.plot(loss_store)
plt.savefig(args.out+'flows_loss_seed_'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()

plt.plot(loss_store_vi)
plt.savefig(args.out+'vi_loss_seed_'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()

sns.kdeplot(sse_betas_mcmc,color='red',label='MCMC')
sns.kdeplot(sse_betas_vi,color='blue',label='MF-VI')
sns.kdeplot(sse_betas_flows,color='green',label='Flows')
plt.legend(fontsize=15)
plt.ylabel('Density of '+r'$||\beta-\beta_{0}||_{2}^{2}$',fontsize=20)
plt.yticks(fontsize=10)
plt.xticks(fontsize=10)
plt.savefig(args.out+'sse_beta_dist'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.pdf')
plt.clf()

sns.kdeplot(acc_samples_mcmc,color='red',label='MCMC')
sns.kdeplot(acc_samples_vi,color='blue',label='MF-VI')
sns.kdeplot(acc_samples_flows,color='green',label='Flows')
plt.legend()
plt.ylabel('Density of Model Accuracy')
plt.tight_layout()
plt.savefig(args.out+'acc_ypred_dist'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()

fig = plt.figure()       
ax = fig.add_subplot(1,2,1)
ax.plot(beta_samples_mh_store[:,0])
ax = fig.add_subplot(1,2,2)
ax.plot(beta_samples_mh_store[:,1])
plt.savefig(args.out+'MH_Diagnostics/trace_seed'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()
mh_samples = pd.DataFrame(beta_samples_mh_store)
fig, ax = plt.subplots(1,2)       
sm.graphics.tsa.plot_acf(mh_samples.iloc[:,0], lags=5, ax = ax[0])
sm.graphics.tsa.plot_acf(mh_samples.iloc[:,1], lags=5, ax = ax[1])
plt.savefig(args.out+'MH_Diagnostics/auto_corr_seed'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
plt.clf()
    
print("accept",accept)
print("Acc",acc)
print("Time in s",Time)

#####################################################INFERENCE##########################################################################

q_lower_mcmc = np.quantile(beta_samples_mh_store,0.025,axis=0)
q_upper_mcmc = np.quantile(beta_samples_mh_store,0.975,axis=0)
q_lower_vi = np.quantile(beta_samples_VI,0.025,axis=0)
q_upper_vi = np.quantile(beta_samples_VI,0.975,axis=0)
q_lower_flows = np.quantile(data[:,0:p],0.025,axis=0)
q_upper_flows = np.quantile(data[:,0:p],0.975,axis=0)

#Checking if true beta0 lies in estimated CI
cint_mcmc = (q_lower_mcmc<= beta0) & (beta0 <= q_upper_mcmc) 
cint_vi = (q_lower_vi<= beta0) & (beta0 <= q_upper_vi)
cint_flows = (q_lower_flows<= beta0) & (beta0 <= q_upper_flows)
correct_pct = {"MH":None,"Flows":None,"MF-VI":None}
correct_pct["MH"] = sum(cint_mcmc)/(cint_mcmc).shape[0]
correct_pct["MF-VI"] = sum(cint_vi)/(cint_vi).shape[0]
correct_pct["Flows"]= sum(cint_flows)/(cint_flows).shape[0]

#Testing null hypothesis beta0 = 0
ht_mcmc = (q_lower_mcmc> np.zeros(p)) | (np.zeros(p) > q_upper_mcmc) 
ht_vi = (q_lower_vi > np.zeros(p)) | (np.zeros(p) > q_upper_vi)
ht_flows = (q_lower_flows > np.zeros(p)) | (np.zeros(p) > q_upper_flows)
ht_beta0 = beta0!=0
confusion = {"MH":None,"Flows":None,"MF-VI":None}
confusion["MH"] = metrics.confusion_matrix(ht_beta0,ht_mcmc)
confusion["MF-VI"] = metrics.confusion_matrix(ht_beta0,ht_vi)
confusion["Flows"] = metrics.confusion_matrix(ht_beta0,ht_flows)

##Calculating F-score
prec = np.array([sum(ht_beta0*ht_mcmc)/sum(ht_mcmc),sum(ht_beta0*ht_flows)/sum(ht_flows),sum(ht_beta0*ht_vi)/sum(ht_vi)])
recall = np.array([sum(ht_beta0*ht_mcmc)/sum(ht_beta0),sum(ht_beta0*ht_flows)/sum(ht_beta0),sum(ht_beta0*ht_vi)/sum(ht_beta0)])

fscore_l=(2*prec*recall)/(prec+recall)
fscore_dict = {"MH":fscore_l[0],"Flows":fscore_l[1],"MF-VI":fscore_l[2]}


print("Correct prop",correct_pct)
print("Confusion MCMC",confusion["MH"])
print("Confusion MF-VI",confusion["MF-VI"])
print("Confusion Flows",confusion["Flows"])
print("Fscore",fscore_dict)
print("SE_min",se_beta_min)
print("SE_avg",se_beta_avg)
print("SE_max",se_beta_max)

###Further Diagnostic Analysis. This is not required in general to reproduce results in paper.
#if args.seed==3:
#    print("beta0",beta0)
#    beta_nonzero=beta0.nonzero()
#    for i in range(beta_nonzero[0].shape[0]):
#        idxb = beta_nonzero[0][i]
#        plt.figure(figsize=(6.4,4.8))
#        sns.kdeplot(beta_samples_mh_store[:,idxb],color='red',label='MCMC')
#        sns.kdeplot(beta_samples_VI[:,idxb],color='blue',label='MF-VI')
#        sns.kdeplot(data[:,idxb],color='green',label='Flows')
#        plt.axvline(x=beta0[idxb], color='gray',ls='--', label='Truth')
#        plt.axvline(x=0, color='purple',ls='--')
#        plt.axvline(x=q_lower_vi[idxb], color='blue',ls='--',label=r'$q_{lower}$'+' MFVI')
#        plt.axvline(x=q_lower_mcmc[idxb], color='red',ls='--',label=r'$q_{lower}$'+' MCMC')
#        plt.axvline(x=q_lower_flows[idxb], color='green',ls='--',label=r'$q_{lower}$'+' Flows')
#        plt.xlabel(r'$\beta$'+'_'+str(idxb))
#        plt.legend()   
#        plt.tight_layout()
#        plt.savefig(args.out+'Diagnostics/'+'beta_idx'+str(idxb)+'trial'+str(args.seed)+'_ndata_'+str(args.n_data)+'_p_'+str(args.data_dim)+'.png')
#        plt.clf()


#############Write Results to Excel######################################################

if args.writecsv==True:
    grid_to_col = {'2_50':0,'2_100':1, '2_200':2, '20_50':3,'20_100':4,'20_200':5,
    '50_50':6,'50_100':7,'50_200':8,'100_50':9,'100_100':10,'100_200':11}
    key = grid_to_col[str(args.data_dim)+'_'+str(args.n_data)]
    from openpyxl import workbook 
    from openpyxl import load_workbook
    filepath = args.out+"Result_logistic.xlsx"
    wb = load_workbook(filepath)
    sheets = wb.sheetnames
    time_sheet = wb['Time']
    time_sheet.cell(row = args.seed+1, column = key+2).value = Time["MH"]
    time_sheet.cell(row = args.seed+1 + 5 , column = key+2).value = Time["Flows"]
    time_sheet.cell(row =args.seed+1 + (5*2), column = key+2).value = Time["MF-VI"]
    acc_sheet = wb['Acc']
    acc_sheet.cell(row = args.seed+1 , column = key+2).value = acc["MH"]
    acc_sheet.cell(row = args.seed+1 + 5, column = key+2).value = acc["Flows"]
    acc_sheet.cell(row = args.seed+1 + (5*2), column = key+2).value = acc["MF-VI"]
    se_betas = wb['SE_betas']
    se_betas.cell(row =args.seed+1, column = key+2).value = se_beta_avg["MH"]
    se_betas.cell(row =args.seed+1 +5, column = key+2).value = se_beta_avg["Flows"]
    se_betas.cell(row =args.seed+1 + 10, column = key+2).value = se_beta_avg["MF-VI"]    
    fscore_s = wb['Fscore']
    fscore_s.cell(row =args.seed+1, column = key+2).value = fscore_dict["MH"]
    fscore_s.cell(row =args.seed+1 +5, column = key+2).value = fscore_dict["Flows"]
    fscore_s.cell(row =args.seed+1 + 10, column = key+2).value = fscore_dict["MF-VI"]        
    cpct = wb['Correct_pct']
    cpct.cell(row =args.seed+1, column = key+2).value = correct_pct["MH"]
    cpct.cell(row =args.seed+1 +5, column = key+2).value = correct_pct["Flows"]
    cpct.cell(row =args.seed+1 + 10, column = key+2).value = correct_pct["MF-VI"]    
    wb.save(filepath) 


